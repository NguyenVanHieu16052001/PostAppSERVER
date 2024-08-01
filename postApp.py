from flask import Flask, jsonify, request
import openpyxl
from datetime import datetime
from openpyxl.styles import Alignment, NamedStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.platypus import Spacer
import subprocess
import os
import signal
from flask_cors import CORS, cross_origin

TEN_SHOP = "SHOP NGUYEN TRUNG" # Tên shop
san_pham_dict = {}
cart_data = {}
cart_data_hang_tra = {}


def doc_file_excel(file_path):
    global san_pham_dict
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua dòng tiêu đề
            ma_sp = row[0]
            ma_sp = str(ma_sp).strip()
            gia_ban_str = str(row[8])
            # Chuyển đổi giá bán từ chuỗi sang số, loại bỏ dấu phẩy
            gia_ban_str = gia_ban_str.replace(',', '')
            gia_ban = int(float(gia_ban_str))

            # Thêm sản phẩm vào từ điển
            san_pham_dict[ma_sp] = gia_ban

    except Exception as e:
        print(f"ERR read file Excel: {e}")

    return san_pham_dict

# Sử dụng hàm
file_path = "HangHoa.xlsx"  # Thay thế bằng đường dẫn thực tế
san_pham_dict = doc_file_excel(file_path)
if(len(san_pham_dict) !=0 ):
    app = Flask(__name__)
    CORS(app)
    app.config['CORS_HEADERS'] = 'Content-Type'
    @app.route('/api/cart', methods=['GET'])
    @cross_origin()
    def get_cart():
        return jsonify(cart_data, cart_data_hang_tra)
    
    @app.route('/api/cart/add/<cart_id>', methods=['POST'])
    @cross_origin()
    def add_to_cart(cart_id):
        data = request.get_json()
        product_id, quantity, price = data['product_id'], data['quantity'], data['price']
        product_id = str(product_id)
        if price == 0:
            try:
                price = san_pham_dict.get(product_id)
                if price == None:
                    return jsonify({"error": "Product not found"})
            except Exception as e:
                return jsonify({"error": e})
        print({product_id, quantity, price})
        if int(cart_id ) == 1:
            if product_id == "2":
                for key, value in cart_data.items():
                    if "New" in str(key) and value[1] == price:
                        cart_data[key][0] += quantity
                        return jsonify(cart_data, cart_data_hang_tra)
                ngay_gio_hien_tai = datetime.now()
                newKey = "New" + ngay_gio_hien_tai.strftime("%H_%M_%S")
                cart_data[newKey] = [quantity, price]
                return jsonify(cart_data, cart_data_hang_tra)
            if product_id in cart_data:
                cart_data[product_id][0] += quantity
            else:
                cart_data[product_id] = [quantity, price]
        elif int(cart_id ) == 2:
            if product_id == "2":
                for key, value in cart_data_hang_tra.items():
                    if "New" in key and value[1] == price:
                        cart_data_hang_tra[key][0] += quantity
                        return jsonify(cart_data, cart_data_hang_tra)
                ngay_gio_hien_tai = datetime.now()
                newKey = "New" + ngay_gio_hien_tai.strftime("%H_%M_%S")
                cart_data_hang_tra[newKey] = [quantity, price]
                return jsonify(cart_data, cart_data_hang_tra)
            if product_id in cart_data_hang_tra:
                cart_data_hang_tra[product_id][0] += quantity
            else:
                cart_data_hang_tra[product_id] = [quantity, price]
        return jsonify(cart_data, cart_data_hang_tra)


    @app.route('/api/cart/remove/<product_id>/<cart_id>', methods=['DELETE'])
    @cross_origin()
    def remove_from_cart(product_id, cart_id):
        product_id = str(product_id)
        if int(cart_id ) == 1:
            if product_id in cart_data:
                del cart_data[product_id]
            else:
                return jsonify({"error": "Product not found in cart"})
        elif int(cart_id ) == 2 or int(cart_id ) == "2":
            if product_id in cart_data_hang_tra:
                del cart_data_hang_tra[product_id]
            else:
                return jsonify({"error": "Product not found in cart"})
        return jsonify(cart_data, cart_data_hang_tra)

    @app.route('/api/cart/update/<product_id>/<cart_id>/<number>', methods=['PUT'])
    @cross_origin()
    def decrease_quantity(product_id, cart_id, number):
        product_id = str(product_id)
        if int(cart_id ) == 1:
            if product_id in cart_data:
                cart_data[product_id][0] = int(number)
            else:
                return jsonify({"error": "Product not found in cart"})
        elif int(cart_id ) == 2 or int(cart_id ) == "2":
            if product_id in cart_data_hang_tra:
                cart_data_hang_tra[product_id][0] = int(number)
            else:
                return jsonify({"error": "Product not found in cart"})
        return jsonify(cart_data, cart_data_hang_tra)
            
    @app.route('/api/<check_printer>', methods=['POST'])
    @cross_origin()
    def tao_hoa_don_2(check_printer):
        global san_pham_dict
        global centered_style_new
        global cart_data_hang_tra
        global cart_data
        ngay_gio_hien_tai = datetime.now()
        ten_file = ngay_gio_hien_tai.strftime("%d_%m_%Y") + ".xlsx"
        hoa_don_pdf_file = ngay_gio_hien_tai.strftime("%d_%m_%Y_%H_%M_%S") + ".pdf"
        ten_sheet = ngay_gio_hien_tai.strftime("%H") + "H"

        try:
            workbook = openpyxl.load_workbook(ten_file)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        if "centered_style" in workbook.named_styles:
            centered_style_new = workbook.named_styles[len(workbook.named_styles) -1]
        else:
            # Tạo kiểu định dạng căn giữa
            centered_style_new = NamedStyle(name="centered_style")
            centered_style_new.alignment = Alignment(horizontal='center', vertical='center')

        check = False
        check_doanh_thu = False
        if "DoanhThu" not in workbook.sheetnames:
            workbook.create_sheet("DoanhThu")
            check_doanh_thu = True
        sheet_doanh_thu = workbook["DoanhThu"]
        if check_doanh_thu:
            sheet_doanh_thu.cell(row=1, column=1, value=0)
        if ten_sheet not in workbook.sheetnames:
            workbook.create_sheet(ten_sheet)
            check = True
        sheet = workbook[ten_sheet]
        if check:
            for col_idx in range(1, 9):
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                sheet.column_dimensions[column_letter].auto_size = True
        sheet.cell(row=1, column=1, value="Thời gian").style = centered_style_new
        sheet.cell(row=1, column=2, value="Số tiền thanh toán").style = centered_style_new
        sheet.cell(row=1, column=3, value="Tiền KM").style = centered_style_new
        sheet.cell(row=1, column=4, value="Mua/Trả").style = centered_style_new
        sheet.cell(row=1, column=5, value="Mã SP").style = centered_style_new
        sheet.cell(row=1, column=6, value="Tên SP").style = centered_style_new
        sheet.cell(row=1, column=7, value="Số lượng").style = centered_style_new
        sheet.cell(row=1, column=8, value="Đơn giá").style = centered_style_new

        try:
            data = request.get_json()
            san_pham_mua = cart_data
            so_tien_giam = data['so_tien_giam']
            san_pham_tra = cart_data_hang_tra

            tong_tien = 0
            for ma_sp, so_luong_don_gia in san_pham_mua.items():
                tong_tien += so_luong_don_gia[0] * so_luong_don_gia[1]

            for ma_sp, so_luong_don_gia in san_pham_tra.items():
                tong_tien -= so_luong_don_gia[0] * so_luong_don_gia[1]

            tong_sp = len(san_pham_mua) + len(san_pham_tra)

            # Thêm dòng mới
            row_index = sheet.max_row + 1
            sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index+tong_sp -1, end_column=1)
            sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index+tong_sp -1, end_column=2)
            sheet.merge_cells(start_row=row_index, start_column=3, end_row=row_index+tong_sp -1, end_column=3)
            if len(san_pham_mua) == 0:
                sheet.merge_cells(start_row=row_index, start_column=4, end_row=row_index+len(san_pham_mua), end_column=4)
            else:
                sheet.merge_cells(start_row=row_index, start_column=4, end_row=row_index+len(san_pham_mua) -1, end_column=4)
            if len(san_pham_tra) == 0:
                sheet.merge_cells(start_row=row_index+len(san_pham_mua), start_column=4, end_row=row_index+tong_sp , end_column=4)
            else:    
                sheet.merge_cells(start_row=row_index+len(san_pham_mua), start_column=4, end_row=row_index+tong_sp -1, end_column=4)
            sheet.cell(row=row_index, column=1, value=ngay_gio_hien_tai.strftime("%H:%M:%S")).style = centered_style_new
            tong_tien_formatted = "{:,.0f} VND".format(tong_tien-so_tien_giam)
            sheet.cell(row=row_index, column=2, value=tong_tien_formatted).style = centered_style_new
            doanh_thu_hien_tai = sheet_doanh_thu.cell(row=1, column=1).value
            tong_doanh_thu = doanh_thu_hien_tai + (tong_tien-so_tien_giam)
            sheet_doanh_thu.cell(row=1, column=1).value = tong_doanh_thu
            so_tien_giam_formatted = "Note: Đã giảm {:,.0f} VND".format(so_tien_giam)
            sheet.cell(row=row_index, column=3, value=so_tien_giam_formatted).style = centered_style_new

            # # Cột (Mua/Trả)
            sheet.cell(row=row_index, column=4, value="Mua").style = centered_style_new
            sheet.cell(row=row_index+len(san_pham_mua), column=4, value="Trả").style = centered_style_new

            # Tiêu đề bảng
            data = [["Ten SP", "SL", "Gia", "Thanh tien"]]
            # # Các cột chi tiết sản phẩm (Số lượng, Đơn giá)
            for ma_sp, so_luong_don_gia in san_pham_mua.items():
                sheet.cell(row=row_index, column=5, value=ma_sp).style = centered_style_new
                sheet.cell(row=row_index, column=6, value="SHOP NGUYEN TRUNG").style = centered_style_new
                sheet.cell(row=row_index, column=7, value=so_luong_don_gia[0]).style = centered_style_new
                
                gia_ban = so_luong_don_gia[1]
                if(gia_ban is None):
                    return jsonify(error="Không tìm thấy sản phẩm với mã sản phẩm này."), 404 
                gia_ban_formatted = "{:,.0f} VND".format(gia_ban)
                sheet.cell(row=row_index, column=8, value=gia_ban_formatted).style = centered_style_new
                row_index += 1
                data.append(["SHOP NGUYEN TRUNG", so_luong_don_gia[0], "{:,.0f}".format(gia_ban), "{:,.0f}".format(so_luong_don_gia[0]*gia_ban)])
            data.append([])
            for ma_sp, so_luong_don_gia in san_pham_tra.items():
                sheet.cell(row=row_index, column=5, value=ma_sp).style = centered_style_new
                sheet.cell(row=row_index, column=6, value="SHOP NGUYEN TRUNG").style = centered_style_new
                sheet.cell(row=row_index, column=7, value=so_luong_don_gia[0]).style = centered_style_new
                
                gia_ban = so_luong_don_gia[1]
                if(gia_ban is None):
                    return jsonify(error="Không tìm thấy sản phẩm với mã sản phẩm này."), 404 
                gia_ban_formatted = "{:,.0f} VND".format(gia_ban)
                sheet.cell(row=row_index, column=8, value=gia_ban_formatted).style = centered_style_new
                row_index += 1
                data.append([ "SHOP NGUYEN TRUNG", so_luong_don_gia[0], "- {:,.0f}".format(gia_ban), "- {:,.0f}".format(so_luong_don_gia[0]*gia_ban)])
            data.append([])
            data.append(["", "", "Tong:", "{:,.0f}".format(tong_tien)])
            data.append(["", "", "Giam gia:", "{:,.0f}".format(so_tien_giam)])
            data.append(["", "", "Thanh tien:", "{:,.0f}".format(tong_tien-so_tien_giam)])
            workbook.save(ten_file)
            if check_printer == "false":
                cart_data = {}
                cart_data_hang_tra = {}
                return jsonify(data="Luu excel thanh cong"), 200
            # Tính toán chiều cao ước lượng của hóa đơn
            row_height = 20  # Chiều cao ước tính của mỗi dòng (điều chỉnh nếu cần)
            header_height = 60  # Chiều cao của tiêu đề và ngày giờ
            footer_height = 60  # Chiều cao của tổng, giảm giá, thành tiền và note
            table_height = (len(san_pham_mua) + len(san_pham_tra) + 3 + 2) * row_height  # Chiều cao của bảng
            estimated_height = header_height + table_height + footer_height + 40
            doc = SimpleDocTemplate(
                hoa_don_pdf_file, 
                pagesize=(80 * mm, estimated_height),
                leftMargin=0,  #   lề trái
                rightMargin=0,  #   lề phải
                topMargin=0,   #   lề trên
                bottomMargin=0  #   lề dưới
            )

            # Tạo nội dung hóa đơn
            story = []

            # Tên shop (in đậm, size lớn, căn giữa)
            title_data = [["SHOP NGUYEN TRUNG"]]
            title_table = Table(title_data, colWidths=[80 * mm])
            title_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 14)  # Điều chỉnh kích thước chữ ở đây
            ]))
            story.append(title_table)

            # Ngày giờ (căn giữa)
            time_data = [[ngay_gio_hien_tai.strftime("%d/%m/%Y %H:%M:%S")]]
            time_data = Table(time_data, colWidths=[80 * mm])
            time_data.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 10)  # Điều chỉnh kích thước chữ ở đây
            ]))
            story.append(time_data)

            story.append(Spacer(1, 10))

            # Tạo bảng
            table = Table(data)
            table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('SIZE', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 1)
            ]))
            story.append(table)
            story.append(Spacer(1, 10))

            # Ngày giờ (căn giữa)
            note_Data = [["Thank you\nHang moi duoc doi tra trong vong 7 ngay"]]
            note_Data = Table(note_Data, colWidths=[80 * mm])
            note_Data.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 10)  # Điều chỉnh kích thước chữ ở đây
            ]))
            story.append(note_Data)

            # Build PDF
            doc.build(story)

            command_print(hoa_don_pdf_file)
            try:
                os.remove(hoa_don_pdf_file)
                print(f"File {hoa_don_pdf_file} in và xoa thanh cong.")
                cart_data = {}
                cart_data_hang_tra = {}
                return jsonify({'success': 'File đã được in và xóa thành công.'}), 200
            except OSError as e:
                return jsonify({'error': 'Không thể xóa file'}), 400
        except KeyError:
            return jsonify({'error': 'Dữ liệu không hợp lệ'}), 400

    @app.route('/api', methods=['POST'])
    @cross_origin()
    def tao_hoa_don():
        global san_pham_dict
        global centered_style_new
        ngay_gio_hien_tai = datetime.now()
        ten_file = ngay_gio_hien_tai.strftime("%d_%m_%Y") + ".xlsx"
        hoa_don_pdf_file = ngay_gio_hien_tai.strftime("%d_%m_%Y_%H_%M_%S") + ".pdf"
        ten_sheet = ngay_gio_hien_tai.strftime("%H") + "H"

        try:
            workbook = openpyxl.load_workbook(ten_file)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        if "centered_style" in workbook.named_styles:
            centered_style_new = workbook.named_styles[len(workbook.named_styles) -1]
        else:
            # Tạo kiểu định dạng căn giữa
            centered_style_new = NamedStyle(name="centered_style")
            centered_style_new.alignment = Alignment(horizontal='center', vertical='center')

        check = False
        check_doanh_thu = False
        if "DoanhThu" not in workbook.sheetnames:
            workbook.create_sheet("DoanhThu")
            check_doanh_thu = True
        sheet_doanh_thu = workbook["DoanhThu"]
        if check_doanh_thu:
            sheet_doanh_thu.cell(row=1, column=1, value=0)
        if ten_sheet not in workbook.sheetnames:
            workbook.create_sheet(ten_sheet)
            check = True
        sheet = workbook[ten_sheet]
        if check:
            for col_idx in range(1, 9):
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                sheet.column_dimensions[column_letter].auto_size = True
        sheet.cell(row=1, column=1, value="Thời gian").style = centered_style_new
        sheet.cell(row=1, column=2, value="Số tiền thanh toán").style = centered_style_new
        sheet.cell(row=1, column=3, value="Tiền KM").style = centered_style_new
        sheet.cell(row=1, column=4, value="Mua/Trả").style = centered_style_new
        sheet.cell(row=1, column=5, value="Mã SP").style = centered_style_new
        sheet.cell(row=1, column=6, value="Tên SP").style = centered_style_new
        sheet.cell(row=1, column=7, value="Số lượng").style = centered_style_new
        sheet.cell(row=1, column=8, value="Đơn giá").style = centered_style_new

        try:
            data = request.get_json()
            san_pham_mua = data['san_pham_mua']
            so_tien_giam = data['so_tien_giam']
            san_pham_tra = data.get('san_pham_tra', {})  # Nếu không có sản phẩm trả, mặc định là từ điển rỗng
            checkPrint = data['checkPrint']
            tong_tien = 0
            for ma_sp, so_luong_don_gia in san_pham_mua.items():
                tong_tien += so_luong_don_gia[0] * so_luong_don_gia[1]

            for ma_sp, so_luong_don_gia in san_pham_tra.items():
                tong_tien += so_luong_don_gia[0] * so_luong_don_gia[1]

            tong_sp = len(san_pham_mua) + len(san_pham_tra)

            # Thêm dòng mới
            row_index = sheet.max_row + 1
            sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index+tong_sp -1, end_column=1)
            sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index+tong_sp -1, end_column=2)
            sheet.merge_cells(start_row=row_index, start_column=3, end_row=row_index+tong_sp -1, end_column=3)
            if len(san_pham_mua) == 0:
                sheet.merge_cells(start_row=row_index, start_column=4, end_row=row_index+len(san_pham_mua), end_column=4)
            else:
                sheet.merge_cells(start_row=row_index, start_column=4, end_row=row_index+len(san_pham_mua) -1, end_column=4)
            if len(san_pham_tra) == 0:
                sheet.merge_cells(start_row=row_index+len(san_pham_mua), start_column=4, end_row=row_index+tong_sp , end_column=4)
            else:    
                sheet.merge_cells(start_row=row_index+len(san_pham_mua), start_column=4, end_row=row_index+tong_sp -1, end_column=4)
            sheet.cell(row=row_index, column=1, value=ngay_gio_hien_tai.strftime("%H:%M:%S")).style = centered_style_new
            tong_tien_formatted = "{:,.0f} VND".format(tong_tien-so_tien_giam)
            sheet.cell(row=row_index, column=2, value=tong_tien_formatted).style = centered_style_new
            doanh_thu_hien_tai = sheet_doanh_thu.cell(row=1, column=1).value
            tong_doanh_thu = doanh_thu_hien_tai + (tong_tien-so_tien_giam)
            sheet_doanh_thu.cell(row=1, column=1).value = tong_doanh_thu
            so_tien_giam_formatted = "Note: Đã giảm {:,.0f} VND".format(so_tien_giam)
            sheet.cell(row=row_index, column=3, value=so_tien_giam_formatted).style = centered_style_new

            # # Cột (Mua/Trả)
            sheet.cell(row=row_index, column=4, value="Mua").style = centered_style_new
            sheet.cell(row=row_index+len(san_pham_mua), column=4, value="Trả").style = centered_style_new

            # Tiêu đề bảng
            data = [["Ten SP", "SL", "Gia", "Thanh tien"]]
            # # Các cột chi tiết sản phẩm (Số lượng, Đơn giá)
            for ma_sp, so_luong_don_gia in san_pham_mua.items():
                sheet.cell(row=row_index, column=5, value=ma_sp).style = centered_style_new
                sheet.cell(row=row_index, column=6, value="SHOP NGUYEN TRUNG").style = centered_style_new
                sheet.cell(row=row_index, column=7, value=so_luong_don_gia[0]).style = centered_style_new
                
                gia_ban = so_luong_don_gia[1]
                if(gia_ban is None):
                    return jsonify(error="Không tìm thấy sản phẩm với mã sản phẩm này."), 404 
                gia_ban_formatted = "{:,.0f} VND".format(gia_ban)
                sheet.cell(row=row_index, column=8, value=gia_ban_formatted).style = centered_style_new
                row_index += 1
                data.append(["SHOP NGUYEN TRUNG", so_luong_don_gia[0], "{:,.0f}".format(gia_ban), "{:,.0f}".format(so_luong_don_gia[0]*gia_ban)])
            data.append([])
            for ma_sp, so_luong_don_gia in san_pham_tra.items():
                sheet.cell(row=row_index, column=5, value=ma_sp).style = centered_style_new
                sheet.cell(row=row_index, column=6, value="SHOP NGUYEN TRUNG").style = centered_style_new
                sheet.cell(row=row_index, column=7, value=so_luong_don_gia[0]).style = centered_style_new
                
                gia_ban = so_luong_don_gia[1]
                if(gia_ban is None):
                    return jsonify(error="Không tìm thấy sản phẩm với mã sản phẩm này."), 404 
                gia_ban_formatted = "{:,.0f} VND".format(gia_ban)
                sheet.cell(row=row_index, column=8, value=gia_ban_formatted).style = centered_style_new
                row_index += 1
                data.append([ "SHOP NGUYEN TRUNG", so_luong_don_gia[0], "- {:,.0f}".format(gia_ban), "- {:,.0f}".format(so_luong_don_gia[0]*gia_ban)])
            data.append([])
            data.append(["", "", "Tong:", "{:,.0f}".format(tong_tien)])
            data.append(["", "", "Giam gia:", "{:,.0f}".format(so_tien_giam)])
            data.append(["", "", "Thanh tien:", "{:,.0f}".format(tong_tien-so_tien_giam)])
            workbook.save(ten_file)
            if checkPrint == "0" or checkPrint == 0:
                return jsonify("ok")
            # Tính toán chiều cao ước lượng của hóa đơn
            row_height = 20  # Chiều cao ước tính của mỗi dòng (điều chỉnh nếu cần)
            header_height = 60  # Chiều cao của tiêu đề và ngày giờ
            footer_height = 60  # Chiều cao của tổng, giảm giá, thành tiền và note
            table_height = (len(san_pham_mua) + len(san_pham_tra) + 3 + 2) * row_height  # Chiều cao của bảng
            estimated_height = header_height + table_height + footer_height + 40
            doc = SimpleDocTemplate(
                hoa_don_pdf_file, 
                pagesize=(80 * mm, estimated_height),
                leftMargin=0,  #   lề trái
                rightMargin=0,  #   lề phải
                topMargin=0,   #   lề trên
                bottomMargin=0  #   lề dưới
            )

            # Tạo nội dung hóa đơn
            story = []

            # Tên shop (in đậm, size lớn, căn giữa)
            title_data = [["SHOP NGUYEN TRUNG"]]
            title_table = Table(title_data, colWidths=[80 * mm])
            title_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 14)  # Điều chỉnh kích thước chữ ở đây
            ]))
            story.append(title_table)

            # Ngày giờ (căn giữa)
            time_data = [[ngay_gio_hien_tai.strftime("%d/%m/%Y %H:%M:%S")]]
            time_data = Table(time_data, colWidths=[80 * mm])
            time_data.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 10)  # Điều chỉnh kích thước chữ ở đây
            ]))
            story.append(time_data)

            story.append(Spacer(1, 10))

            # Tạo bảng
            table = Table(data)
            table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('SIZE', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 1)
            ]))
            story.append(table)
            story.append(Spacer(1, 10))

            # Ngày giờ (căn giữa)
            note_Data = [["Thank you\nHang moi duoc doi tra trong vong 7 ngay"]]
            note_Data = Table(note_Data, colWidths=[80 * mm])
            note_Data.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 10)  # Điều chỉnh kích thước chữ ở đây
            ]))
            story.append(note_Data)

            # Build PDF
            doc.build(story)

            command_print(hoa_don_pdf_file)
            try:
                os.remove(hoa_don_pdf_file)
                print(f"File {hoa_don_pdf_file} in và xoa thanh cong.")
                return jsonify({'success': 'File đã được in và xóa thành công.'}), 200
            except OSError as e:
                return jsonify({'error': 'Không thể xóa file'}), 400
        except KeyError:
            return jsonify({'error': 'Dữ liệu không hợp lệ'}), 400
        
    @app.route('/')
    @cross_origin()
    def hello():
        return "Xin chào! Đây là máy chủ web của Anh Hiếu! Heheheeeeeeeee"

    @app.route('/api/<int:product_id>', methods=['GET'])
    @cross_origin()
    def getPrice(product_id):
        gia = san_pham_dict.get(product_id)
        if gia is not None:
            return jsonify(gia=gia)
        else:
            return jsonify(error="Không tìm thấy sản phẩm với mã sản phẩm này."), 404
    def command_print(hoa_don_pdf_file, event = None):
                command = "{} {}".format('PDFtoPrinter.exe',hoa_don_pdf_file)
                subprocess.call(command,shell=True)
    def clearData():
        global cart_data
        global cart_data_hang_tra
        cart_data = {}
        cart_data_hang_tra = {}
    @app.route('/api/cancel', methods=['DELETE'])
    @cross_origin()
    def cancel():
        clearData()
        return jsonify("ok")
    
def shutdown_server():
    os.system(f"fuser -k 8080/tcp")  # Hoặc sử dụng lsof + kill

signal.signal(signal.SIGTERM, shutdown_server)
if __name__ == '__main__':
    app.run(host='192.168.1.99', port=8080) 